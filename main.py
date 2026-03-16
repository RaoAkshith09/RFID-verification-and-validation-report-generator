import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import pandas as pd
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl import load_workbook
import numpy as np
import warnings
pd.options.mode.chained_assignment = None
warnings.filterwarnings("ignore")

#---------------------------
#-------------report heading--------------
#---------------------------
from openpyxl.styles import Alignment, Font
import pandas as pd

def write_report_with_header(df, output_path,division):
    section = df['Section name'].iloc[0]

    if df['Loco Direction'].iloc[0] == "Nominal":
        line = "DN LINE (NOMINAL)"
    else:
        line = "UP LINE (REVERSE)"
    df = df.reset_index(drop=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=7)
        sheet = writer.book.active

        header_text = (
            "HBL Engineering Ltd\n"
            f"Project: Kavach in {division} Division\n"
            "Phase-I RFID Tag Verification Report\n"
            f"{section} Block section\n"
            "Tag Data & Layout version: \n"
            f"Line: {line}"
        )

        max_col = df.shape[1]

        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=6,
            end_column=max_col
        )

        cell = sheet.cell(row=1, column=1)
        cell.value = header_text
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.font = Font(bold=True)

        for r in range(1, 7):
            sheet.row_dimensions[r].height = 14

#---------------------------
#-------------tag encoder--------------
#---------------------------
def encoder(tagread,ref_dir):
    initcol=[
        'Type of the tag(As programmed)', 'Stnid_N(As programmed)','Stnid_R(As programmed)',
     'Section Type','ABS Location (As programmed)','Placement(as per programmed)','prev','Tin-Id(Nominal)','Tin-Id(Reverse)'
    ]
    
    for col in initcol:
        tagread[col] = pd.NA
    tagread['Remarks on missing of the  tags as per the sequence']=''
    for i, row in tagread.iterrows():
        try:
            if pd.isna(row['Tag Data']):
                raise ValueError

            page_x, page_y = row['Tag Data'].split()
            px_bin = bin(int(page_x, 16))[2:].zfill(64)
            py_bin = bin(int(page_y, 16))[2:].zfill(64)

            stnid_N = int(py_bin[57:64] + px_bin[0:9], 2)
            stnid_R = int(py_bin[41:57], 2)
            abs_loc = int(px_bin[25:48], 2)
            dup_bit = int(py_bin[32], 2)
            section_type_N = int(py_bin[39:41], 2)
            section_type_R = int(py_bin[37:39], 2)
            tin_n = int(px_bin[17:25], 2)
            tin_r = int(px_bin[9:17], 2)
            '''if ref_dir=='Nominal' and len(str(stnid_N))==3:
                    tagread.at[i, 'Section Type(N)'] = 'absolute block'
                    tagread.at[i, 'Section Type(R)'] = 'absolute block' if section_type_R == 1 else 'Station'
            elif ref_dir=='Reverse' and len(str(stnid_N))==3:
                tagread.at[i, 'Section Type(R)'] = 'absolute block' 
                tagread.at[i, 'Section Type(N)'] = 'absolute block' if section_type_N == 1 else 'Station'''
            
            tagread.at[i, 'Section Type(N)'] = 'absolute block' if section_type_N == 1 else 'station'
            tagread.at[i, 'Section Type(R)'] = 'absolute block' if section_type_R == 1 else 'station'

            tagread.at[i, 'Stnid_N(As programmed)'] = stnid_N
            tagread.at[i, 'Stnid_R(As programmed)'] = stnid_R
            tagread.at[i, 'ABS Location (As programmed)'] = abs_loc
            tagread.at[i, 'Type of the tag(As programmed)'] = 'D' if dup_bit == 1 else 'M'
            tagread.at[i, 'Tin-Id(Nominal)'] = tin_n
            tagread.at[i, 'Tin-Id(Reverse)']=tin_r



            
            
            tag_placement_num = int(py_bin[33:37], 2)
            if tag_placement_num<10:
                tag_placement = {
                0: "In line section",
                1: "Signal foot (N)",
                2: "Signal foot (R)",
                3: "Turnout",
                4: "Exit (N)",
                5: "Exit (R)",
                6: "Signal foot (B)",
                7: "Exit (B)",
                8: "Dead Stop (N)",
                9: "Dead Stop (R)",
            }.get(tag_placement_num, "Unknown")
            else:
                tag_placement="Reserved"
            tagread.at[i,'Placement(as per programmed)']=tag_placement


        except Exception:
            tagread.at[i, 'TYPE of the tag'] = 'Error'
    tagread[['PAGE-X (as programmed)', 'PAGE-Y (as programmed)']] = (tagread['Tag Data'].astype(str).str.upper().str.split(' ', expand=True))


    tagread.rename(columns={'TagId': 'Tag-Id(As read by loco)'}, inplace=True)
    #tagread = tagread.reset_index(drop=True)
    tagread = tagread.sort_values(by='TimeStamp').reset_index(drop=True)
    tagread = tagread.reset_index(drop=True)
    for i in range(0, len(tagread) - 1, 2):
        tag1 = tagread.loc[i, 'Tag-Id(As read by loco)']
        tag2 = tagread.loc[i + 1, 'Tag-Id(As read by loco)']
        if tag1==tag2:
            diff= tagread.loc[i, 'ABS Location (As programmed)']-tagread.loc[i + 1, 'ABS Location (As programmed)']
            if ref_dir=="Nominal":
                if diff>0:
                    tagread.loc[i + 1, 'Remarks on missing of the  tags as per the sequence']='Inter change'
                    tagread.loc[i, 'Remarks on missing of the  tags as per the sequence']='Inter change'
            else:
                if diff<0:
                    tagread.loc[i + 1, 'Remarks on missing of the  tags as per the sequence']='Inter change'
                    tagread.loc[i, 'Remarks on missing of the  tags as per the sequence']='Inter change'
        

    tagread=tagread[['Tag-Id(As read by loco)','Type of the tag(As programmed)','Placement(as per programmed)','PAGE-X (as programmed)', 'PAGE-Y (as programmed)',
                     'Stnid_N(As programmed)','Stnid_R(As programmed)','ABS Location (As programmed)','Section Type(N)','Section Type(R)','TimeStamp','prev','Tin-Id(Nominal)','Tin-Id(Reverse)','Remarks on missing of the  tags as per the sequence']]
    #pivot.to_excel(f'C:/Users/Akshith/Desktop/new db pro max/sample/1-1-26/New folder/pivot_{ref_dir}.xlsx')
    return tagread

#-------------report gen--------------
#---------------------------
def clean_and_parse(series):

    # If already datetime → return as is
    if pd.api.types.is_datetime64_any_dtype(series):
        return series

    series = series.astype(str).str.strip()

    # Case 1: Already ISO format (2026-02-13 09:19:09.805000)
    iso_mask = series.str.match(r'\d{4}-\d{2}-\d{2}')

    # Parse ISO directly
    series.loc[iso_mask] = pd.to_datetime(
        series.loc[iso_mask],
        errors='coerce'
    )

    # Case 2: Original format HH:MM:SS:MS DD/MM/YYYY
    orig_mask = ~iso_mask

    temp = series.loc[orig_mask]

    temp = temp.str.replace(
        r'(\d{2}:\d{2}:\d{2}):(\d{3})',
        r'\1.\2',
        regex=True
    )

    temp = pd.to_datetime(
        temp,
        format='%H:%M:%S.%f %d/%m/%Y',
        errors='coerce'
    )

    series.loc[orig_mask] = temp

    return pd.to_datetime(series, errors='coerce')







#-------------report gen--------------
#---------------------------
def report_generation(master_data_set,dmi_display_dir,tag_read,ref_dir,main_fodler,output,division,dmi_display_main):
    tag_read = tag_read.drop_duplicates(subset=['TagId', 'Tag Data'])
    master_data_set['STN_TYP_N'] = (master_data_set['STN_TYP_N'].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
        .str.replace(r'(?i)absolute\s*block', 'absolute block', regex=True)
        .str.lower())
    master_data_set['STN_TYP_R'] = (master_data_set['STN_TYP_R'].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
        .str.replace(r'(?i)absolute\s*block', 'absolute block', regex=True)
        .str.lower())
    '''dmi_display_main['TimeStamp'] = pd.to_datetime(
        dmi_display_main['TimeStamp'],
        format='%H:%M:%S:%f %d/%m/%Y',
        errors='coerce'
    )

    dmi_display_dir['TimeStamp'] = pd.to_datetime(
        dmi_display_dir['TimeStamp'],
        format='%H:%M:%S:%f %d/%m/%Y',
        errors='coerce'
    )

    tag_read['TimeStamp'] = pd.to_datetime(
        tag_read['TimeStamp'],
        format='%H:%M:%S:%f %d/%m/%Y',
        errors='coerce'
    )'''
    dmi_display_main.loc[:, 'TimeStamp'] = clean_and_parse(dmi_display_main['TimeStamp'])
    dmi_display_dir.loc[:, 'TimeStamp']  = clean_and_parse(dmi_display_dir['TimeStamp'])
    tag_read.loc[:, 'TimeStamp']         = clean_and_parse(tag_read['TimeStamp'])
    #print(f" time stamp= {dmi_display_main['TimeStamp'].iloc[0]}")
    #print(f" micro secs{dmi_display_main['TimeStamp'].iloc[0].microsecond}")




    #dmi_display_main['TimeStamp'] = dmi_display_main['TimeStamp'].dt.time
    #dmi_display_dir['TimeStamp']  = dmi_display_dir['TimeStamp'].dt.time
    #tag_read['TimeStamp']  = tag_read['TimeStamp'].dt.time

#    tag_read['TimeStamp'] = pd.to_datetime(tag_read['TimeStamp'], format='%H:%M:%S:%f').dt.time
    master_data_set[['TagId(As per design)', 'Type of tag(As per design)']] = (master_data_set['UID'].astype(str).str.split('-', expand=True))
    ref_fst_tag  = dmi_display_dir['TagNo'].iloc[0]
    ref_fst_time = dmi_display_dir['TimeStamp'].iloc[0]
    ref_end_tag  = dmi_display_dir['TagNo'].iloc[-1]
    ref_end_time = dmi_display_dir['TimeStamp'].iloc[-1]
    dmi_display_main=dmi_display_main[(dmi_display_main['TimeStamp']>=ref_fst_time)&(dmi_display_main['TimeStamp']<=ref_end_time)]
    #print(f'reffirst tag={ref_fst_tag}')
    #print(f'reffirst time={ref_fst_time}')
    #print(f'reflast tag={ref_end_tag}')
    #print(f'ref last tag={ref_end_time}')
    #print(dmi_display_main['TimeStamp'].iloc[0])
    tag_read_ref=encoder(tag_read,ref_dir)
    tag_ref=tag_read_ref.copy()
    tag_ref.to_excel(os.path.join(output,f'tagread_ref_{ref_dir}.xlsx'))
    tag_read_ref.to_excel(os.path.join(output,f'tagread_{ref_dir}.xlsx'))
    try:
        ref_fst_idx=tag_read_ref[(tag_read_ref['TimeStamp']>=ref_fst_time)].index[0]
        ref_last_idx=tag_read_ref[(tag_read_ref['TimeStamp']<=ref_end_time)].index[-1]
    

    except:
        ref_fst_idx=tag_read_ref[(tag_read_ref['TimeStamp']<=ref_fst_time)].index[0]
        ref_last_idx=tag_read_ref[(tag_read_ref['TimeStamp']<=ref_end_time)].index[-1]
    
    #print(f"{ref_fst_idx},{ref_last_idx}")
    #print(tag_read_ref['Tag-Id(As read by loco)'].iloc[ref_fst_idx])
    tag_read_ref=tag_read_ref.iloc[ref_fst_idx:ref_last_idx+1]
    
    
    master_data_set["TagId(As per design)"] = (
        pd.to_numeric(
            master_data_set["TagId(As per design)"],
            errors="coerce"
        ).astype("Int64")
    )

    tag_read_ref["Tag-Id(As read by loco)"] = (
        pd.to_numeric(
            tag_read_ref["Tag-Id(As read by loco)"],
            errors="coerce"
        ).astype("Int64")
    )


    if ref_dir=="Nominal":
        #master_data_set=master_data_set[master_data_set['STN_TYP_N']=="absolute block"].copy()
        #tag_read_ref=tag_read_ref[tag_read_ref['Section Type(N)']=='Block'].copy()
        merged = master_data_set.merge(
            tag_read_ref,
            how="left",
            left_on=["TagId(As per design)", "Type of tag(As per design)","TIN_N"],
            right_on=["Tag-Id(As read by loco)", "Type of the tag(As programmed)","Tin-Id(Nominal)"]
        )
        merged=merged[['Dir','Section name','TagId(As per design)','Tag-Id(As read by loco)','Type of tag(As per design)','Type of the tag(As programmed)',
                       'TIN_N','Tin-Id(Nominal)','Placement(as per programmed)','STN_TYP_N','Section Type(N)',
                       'STN_ID_N','Stnid_N(As programmed)','Xdata','PAGE-X (as programmed)','Ydata','PAGE-Y (as programmed)',
                       'ABS_LOC_M','ABS Location (As programmed)','Remarks on missing of the  tags as per the sequence','TimeStamp'
                        ]]
        merged = merged.rename(columns={'STN_ID_N': 'Stnid(As per design)','Stnid_N(As programmed)':'Stnid(As programmed)',
                                        'Tin-Id(Nominal)':'Tin-Id(As programmed)','TIN_N':'Tin-Id(As per design)',
                                        'STN_TYP_N':'Section type(As per design)','Section Type(N)':'Section type(As programmed)'})


        #merged.to_excel(r'C:\Users\Akshith\Desktop\new db pro max\sample\output_N.xlsx')
    else:
        #master_data_set=master_data_set[master_data_set['STN_TYP_R']=="absolute block"].copy()
        #tag_read_ref=tag_read_ref[tag_read_ref['Section Type(R)']=='Block'].copy()
        merged = master_data_set.merge(
            tag_read_ref,
            how="left",
            left_on=["TagId(As per design)", "Type of tag(As per design)","TIN_R"],
            right_on=["Tag-Id(As read by loco)", "Type of the tag(As programmed)","Tin-Id(Reverse)"]
        )
        merged=merged[['Dir','Section name','TagId(As per design)','Tag-Id(As read by loco)','Type of tag(As per design)','Type of the tag(As programmed)',
                       'TIN_R','Tin-Id(Reverse)','Placement(as per programmed)','STN_TYP_R','Section Type(R)',
                       'STN_ID_R','Stnid_R(As programmed)','Xdata','PAGE-X (as programmed)','Ydata','PAGE-Y (as programmed)',
                       'ABS_LOC_M','ABS Location (As programmed)','TimeStamp','Remarks on missing of the  tags as per the sequence'
                        ]]
        merged = merged.rename(columns={'STN_ID_R': 'Stnid(As per design)','Stnid_R(As programmed)':'Stnid(As programmed)',
                                        'Tin-Id(Reverse)':'Tin-Id(As programmed)','TIN_R':'Tin-Id(As per design)',
                                        'STN_TYP_R':'Section type(As per design)','Section Type(R)':'Section type(As programmed)'})
        #merged.to_excel(r'C:\Users\Akshith\Desktop\new db pro max\sample\output_R.xlsx')
    #merged['Remarks on missing of the  tags as per the sequence']=''
    #print(len(merged))
    merged['Remark on tag data']=''
    merged['ABS Location (at which tag read as per loco log)']=pd.NA
    merged['ABS Location (updated after reading the tag as per loco log)']=pd.NA
    merged['ABS correction, due to tag placement']=pd.NA
    merged['Remark on placement']=pd.NA

    for i, row in merged.iterrows():
        if pd.isna(row['Tag-Id(As read by loco)']):
            merged.loc[i, 'Remarks on missing of the  tags as per the sequence'] = "loco missed this tag"
            merged.loc[i,'ABS Location (at which tag read as per loco log)']=1
            merged.loc[i,'ABS Location (updated after reading the tag as per loco log)']=row['ABS_LOC_M']
            merged.loc[i,'ABS Location (As programmed)']=row['ABS_LOC_M']
            #print(f"{row['TagId(As per design)']} | {merged.loc[i,'ABS Location (at which tag read as per loco log)']}")

            
        else:
            
            #disp_rows = dmi_display_dir[(dmi_display_dir['TimeStamp'] >= row['TimeStamp']) &(dmi_display_dir['TagNo'] == row['Tag-Id(As read by loco)'])]
            x_match = str(row['PAGE-X (as programmed)']).lower().strip() == str(row['Xdata']).lower().strip()
            y_match = str(row['PAGE-Y (as programmed)']).lower().strip() == str(row['Ydata']).lower().strip()

            if x_match and y_match:
                remark = 'Matched'
            elif not x_match and y_match:
                remark = 'PAGE-X Mismatch'
            elif x_match and not y_match:
                remark = 'PAGE-Y Mismatch'
            else:
                remark = 'PAGE-X & PAGE-Y Mismatch'

            merged.loc[i, 'Remark on tag data'] = remark
            disp_rows = dmi_display_main[(dmi_display_main['TimeStamp'] > row['TimeStamp']) &(dmi_display_main['TagNo'] == row['Tag-Id(As read by loco)'])]
            bef_loc = pd.NA
            aft_loc = pd.NA
            if disp_rows.empty:
                bef_loc = 0
                aft_loc = pd.NA
            else:
                pos = disp_rows.index[0]
                pos_loc = dmi_display_main.index.get_loc(pos)

                # previous row
                if pos_loc > 0:
                    bef_loc = dmi_display_main.iloc[pos_loc - 1]['AbsLocation']
                else:
                    bef_loc = 0

                # next row
                if pos_loc < len(dmi_display_main) - 1:
                    aft_loc = dmi_display_main.iloc[pos_loc + 1]['AbsLocation']

            merged.loc[i,'ABS Location (updated after reading the tag as per loco log)'] = aft_loc
            merged.loc[i,'ABS Location (at which tag read as per loco log)'] = bef_loc
            merged.loc[i,'ABS correction, due to tag placement'] = (
                bef_loc - row['ABS_LOC_M']
                if pd.notna(bef_loc) and pd.notna(row['ABS_LOC_M'])
                else pd.NA
            )
    merged = merged.rename(columns={'Xdata': 'PAGE-X(As per design)','Ydata': 'PAGE-Y(As per design)','ABS_LOC_M':'ABS Location(As per design)'})
        
    merged=merged[['Dir','Section name','TagId(As per design)','Tag-Id(As read by loco)','Type of tag(As per design)','Type of the tag(As programmed)','Tin-Id(As per design)','Tin-Id(As programmed)',
                   'Placement(as per programmed)','Section type(As per design)','Section type(As programmed)',
                       'Stnid(As per design)','Stnid(As programmed)','PAGE-X(As per design)','PAGE-X (as programmed)','PAGE-Y(As per design)','PAGE-Y (as programmed)',
                       'ABS Location(As per design)','ABS Location (As programmed)','ABS Location (at which tag read as per loco log)','ABS correction, due to tag placement',
                       'Remarks on missing of the  tags as per the sequence','Remark on tag data','Remark on placement','TimeStamp'
                        ]]
    # rechicking of tags 
    #if 'loco missed this tag' in str(row['Remarks on missing of the  tags as per the sequence']):
     #   print("checkinh")
    for index, row in merged.iterrows():

        if row['Remarks on missing of the  tags as per the sequence'] != 'loco missed this tag':

            continue

        sus_tag  = row['TagId(As per design)']
        sus_stn  = row['Tin-Id(As per design)']
        sus_dir  = row['Dir']
        sus_type = row['Type of tag(As per design)']

        # ---- Direction-based column selection ----
        if sus_dir == "Nominal":
            stn_col = 'Tin-Id(Nominal)'
        else:
            stn_col = 'Tin-Id(Reverse)'

        # ---- Filter reference ----
        tag_read_ref_sus = tag_ref[
            (tag_ref['Tag-Id(As read by loco)'] == sus_tag) &
            (tag_ref['Type of the tag(As programmed)'] == sus_type) &
            (tag_ref[stn_col] == sus_stn)
        ]

        if tag_read_ref_sus.empty:
           # print(f"empty → {sus_tag} | {sus_stn} | {sus_dir}")
            continue
        # ---- Extract values ----
        pagex = tag_read_ref_sus['PAGE-X (as programmed)'].iloc[0]
        pagey = tag_read_ref_sus['PAGE-Y (as programmed)'].iloc[0]

        # ---- Update merged dataframe ----
        merged.loc[index, 'Tag-Id(As read by loco)'] = tag_read_ref_sus['Tag-Id(As read by loco)'].iloc[0]
        merged.loc[index, 'Type of the tag(As programmed)'] = tag_read_ref_sus['Type of the tag(As programmed)'].iloc[0]
        merged.loc[index, 'Placement(as per programmed)'] = tag_read_ref_sus['Placement(as per programmed)'].iloc[0]
        merged.loc[index, 'Stnid(As programmed)'] = tag_read_ref_sus[stn_col].iloc[0]
        merged.loc[index, 'PAGE-X (as programmed)'] = pagex
        merged.loc[index, 'PAGE-Y (as programmed)'] = pagey
        merged.loc[index, 'ABS Location (As programmed)'] = tag_read_ref_sus['ABS Location (As programmed)'].iloc[0]
        merged.loc[index, 'TimeStamp'] = tag_read_ref_sus['TimeStamp'].iloc[0]
        merged.loc[index, 'Remarks on missing of the  tags as per the sequence'] = 'Direction undefined'
        merged.loc[index,'ABS Location (at which tag read as per loco log)']=0

        # ---- PAGE comparison ----
        x_match = str(pagex).strip().lower() == str(row['PAGE-X(As per design)']).strip().lower()
        y_match = str(pagey).strip().lower() == str(row['PAGE-Y(As per design)']).strip().lower()

        if x_match and y_match:
            remark = 'Matched'
        elif not x_match and y_match:
            remark = 'PAGE-X Mismatch'
        elif x_match and not y_match:
            remark = 'PAGE-Y Mismatch'
        else:
            remark = 'PAGE-X & PAGE-Y Mismatch'

        merged.loc[index, 'Remark on tag data'] = remark



    base_name=Path(main_fodler).name
    #print(len(merged))
    merged['TimeStamp'] = pd.to_datetime(
        merged['TimeStamp'],
        errors='coerce'
    )

    # Now convert to string with milliseconds
    merged['TimeStamp'] = (
        merged['TimeStamp']
        .dt.strftime('%H:%M:%S.%f %d/%m/%Y')
        .str.replace(r'(\.\d{3})\d{3}', r'\1', regex=True)
    )

    merged.to_excel(os.path.join(output,f"{base_name}_{merged['Dir'].iloc[0]}_Tag_data_validation.xlsx"),index=False)
    

    #merged['Stnid(As per design)'] = merged['Stnid(As programmed)'].astype('Int64')
    for section_name, grp in merged.groupby('Section name'):
        grp = grp.reset_index(drop=True)
        out_grp=grp.copy()
        out_grp = out_grp.reset_index()
        out_grp.rename(columns={'Dir': 'Loco Direction'}, inplace=True)
        mask = (
            out_grp["ABS Location (at which tag read as per loco log)"].notna() &
            (
                (out_grp["ABS Location (at which tag read as per loco log)"] > 8000000) |
                (out_grp["ABS Location (at which tag read as per loco log)"] == 0)
            )
        )

        out_grp.loc[mask, 'Remarks on missing of the  tags as per the sequence'] = 'Direction undefined'
        out_grp = out_grp.drop(columns=['ABS Location(As per design)','ABS Location (at which tag read as per loco log)', 'ABS Location (As programmed)',                                         
                                         'Remark on placement'])
        
        if 'ABS correction, due to tag placement' in out_grp.columns:
            out_grp=out_grp.drop(columns=['ABS correction, due to tag placement'])

        if grp['TimeStamp'].isna().all():
            continue
        tag_map = {
            9: "Normal",
            10: "Gate",
            11: "Adj line",
            12: "Adjustment"
        }

        # Compute tag type directly with apply
        out_grp.insert(
            7,
            "Kind of tag",
            out_grp["PAGE-X(As per design)"].apply(
                lambda x: tag_map.get(
                    int(bin(int(str(x), 16))[2:].zfill(64)[60:64], 2),
                    "Unknown"
                )
            )
        )




        file_path = os.path.join(
            output,
            f"{base_name}_{merged['Dir'].iloc[0]}_{section_name}_Tag_validation_&_verification_report.xlsx"
        )
        write_report_with_header(out_grp, file_path,division)
        grp2=grp.copy()
        tag_sat_report=grp2[['Dir','Section name','TagId(As per design)','Tag-Id(As read by loco)','Type of the tag(As programmed)','Type of tag(As per design)','ABS Location(As per design)','ABS Location (As programmed)','ABS Location (at which tag read as per loco log)']].copy()
        #tag_sat_report=tag_sat_report[tag_sat_report['Type of tag(As per design)']=="M"]
        tag_sat_report = tag_sat_report.drop_duplicates(subset=["TagId(As per design)"])
        tag_sat_report.rename(columns={'Dir': 'Loco Direction'}, inplace=True)
        tag_sat_report['Previous Tag No.']=pd.NA
        tag_sat_report['Tag Data verified as per approved RFID Tag Data (OK/Not OK)']=pd.NA
        tag_sat_report['Distance from previous Tag to current Tag  (As per loco log)']=pd.NA
        tag_sat_report['Distance from previous Tag to current Tag  (As programmed)']=pd.NA
        tag_sat_report['Distance from previous Tag to current Tag  (As per design)']=pd.NA
        tag_sat_report['Allowable error in tag placement in meters (RDSO Spec)  Upper limit']=pd.NA
        tag_sat_report['Allowable error in tag placement in meters (RDSO Spec)  Lower limit']=pd.NA
        tag_sat_report['Remark on placement(Correct/Incorrect))']=pd.NA
        tag_sat_report['Error in Tag Placement \n( + : After  - : Before )']=pd.NA
        tag_sat_report['Corrective Action To be implemented by field Team.']=pd.NA


        tag_sat_report = tag_sat_report.reset_index(drop=True)

        undefined_cols = [
            'ABS Location (at which tag read as per loco log)',
            'Distance from previous Tag to current Tag  (As per design)',
            'Distance from previous Tag to current Tag  (As programmed)',
            'Distance from previous Tag to current Tag  (As per loco log)',
            'Allowable error in tag placement in meters (RDSO Spec)  Upper limit',
            'Allowable error in tag placement in meters (RDSO Spec)  Lower limit',
            'Error in Tag Placement \n( + : After  - : Before )',
            'Remark on placement(Correct/Incorrect))'
        ]

        for i in range(len(tag_sat_report)):

            cur_abs_design = tag_sat_report.loc[i, 'ABS Location(As per design)']
            cur_abs_programmed = tag_sat_report.loc[i, 'ABS Location (As programmed)']
            cur_abs_locoread = tag_sat_report.loc[i, 'ABS Location (at which tag read as per loco log)']
            ct=tag_sat_report.loc[i, 'TagId(As per design)']
            #print(f" {ct} | {cur_abs_locoread} ")
            # skipping the row at tag missed 
            if cur_abs_locoread==1:
                r = grp2.loc[
                    (grp2['TagId(As per design)'] == ct) &
                    (grp2['Type of the tag(As programmed)'] == "D"),
                    'ABS Location (at which tag read as per loco log)'
                ]
                if not r.empty:
                    r = r.iloc[0]
                    if pd.notna(r) and r > 1:
                        tag_sat_report.loc[i, 'ABS Location (at which tag read as per loco log)'] = r
                else:
                    tag_sat_report.loc[i, undefined_cols] = "Loco missed this tag"
                    continue

            # Undefined direction condition
            if (cur_abs_locoread > 7999999) or (cur_abs_locoread == 0):

                tag_sat_report.loc[i, undefined_cols] = "Undefined direction"
                continue  # skip further calculations
            
            # Skip first row (no previous tag)
            if i == 0 and (cur_abs_locoread == 0) :
                tag_sat_report.loc[i, undefined_cols] = "Undefined direction"
                continue

            # Get previous row values
            try:
                if i==0:
                    continue
                prev_row = tag_sat_report.loc[i - 1]

                nxt_tag = prev_row['TagId(As per design)']
                nxt_abs_design = prev_row['ABS Location(As per design)']
                nxt_abs_prog = prev_row['ABS Location (As programmed)']

                # Distance calculations
                design_diff = abs(nxt_abs_design - cur_abs_design)
                prog_diff = abs(nxt_abs_prog - cur_abs_design)
                loco_diff = abs(nxt_abs_design - cur_abs_locoread)

                # RDSO limits
                Upper_limit = abs(design_diff * 1.05 + 5)
                Lower_limit = abs(design_diff * 0.95 - 5)

                error = loco_diff - design_diff

                # Fill values
                tag_sat_report.loc[i, 'Previous Tag No.'] = nxt_tag
                tag_sat_report.loc[i, 'Distance from previous Tag to current Tag  (As per design)'] = design_diff
                tag_sat_report.loc[i, 'Distance from previous Tag to current Tag  (As programmed)'] = prog_diff
                tag_sat_report.loc[i, 'Distance from previous Tag to current Tag  (As per loco log)'] = loco_diff
                tag_sat_report.loc[i, 'Allowable error in tag placement in meters (RDSO Spec)  Upper limit'] = Upper_limit
                tag_sat_report.loc[i, 'Allowable error in tag placement in meters (RDSO Spec)  Lower limit'] = Lower_limit
                tag_sat_report.loc[i, 'Error in Tag Placement \n( + : After  - : Before )'] = error

                # Placement remark
                if Lower_limit <= loco_diff <= Upper_limit:
                    tag_sat_report.loc[i, 'Remark on placement(Correct/Incorrect))'] = "correct"
                else:
                    tag_sat_report.loc[i, 'Remark on placement(Correct/Incorrect))'] = "incorrect"
            except :
                tag_sat_report.loc[i, undefined_cols] = "Error occured"


        # Reorder columns
        tag_sat_report = tag_sat_report[['Loco Direction','Section name','TagId(As per design)','Tag-Id(As read by loco)',
            'Type of tag(As per design)','Type of the tag(As programmed)',            
            'ABS Location(As per design)','ABS Location (As programmed)',
            'ABS Location (at which tag read as per loco log)',
            'Distance from previous Tag to current Tag  (As per design)',
            'Distance from previous Tag to current Tag  (As per loco log)',
            'Error in Tag Placement \n( + : After  - : Before )','Allowable error in tag placement in meters (RDSO Spec)  Upper limit'
            ,'Allowable error in tag placement in meters (RDSO Spec)  Lower limit','Remark on placement(Correct/Incorrect))',
            'Corrective Action To be implemented by field Team.']]

        # Save Excel
        satopp = os.path.join(output,
            f"{base_name}_{tag_sat_report['Loco Direction'].iloc[0]}_{tag_sat_report['Section name'].iloc[0]}_Tag_placement_verification.xlsx")
        #tag_sat_report.to_excel(satopp, index=False)
        # ---------- Rename Page 1 ----------
        wb = load_workbook(file_path)

        # rename first sheet
        wb.worksheets[0].title = "Tag Data Verification"

        wb.save(file_path)

        # ---------- Write DF to Page 2 ----------
        with pd.ExcelWriter(
            file_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            tag_sat_report.to_excel(
                writer,
                sheet_name="Tag Placement Verification",
                index=False
            )


        print(f"file saved: {satopp}")
        


                
        #grp.to_excel(file_path, index=False)

    


    
 # ==============================
# MAIN ENTRY FUNCTION FOR GUI
# ==============================

def run_full_process(stn_main, main_fodler, division):

    stns_folders = [
        f for f in os.listdir(stn_main)
        if os.path.isdir(os.path.join(stn_main, f))
    ]

    for stn in stns_folders:

        master_data = os.path.join(stn_main, stn)
        output = master_data

        print(f"--------------------------{stn}--------------------------")

        tag_read = pd.read_csv(
            next(os.path.join(main_fodler, f)
                 for f in os.listdir(main_fodler)
                 if f.startswith('TagRead'))
        )

        largest_dmi_display_path = next(
            os.path.join(main_fodler, f)
            for f in os.listdir(main_fodler)
            if f.startswith('DmiDisplay')
        )
        try:

            dmi_display = pd.read_csv(largest_dmi_display_path, delimiter=";")
            
        except:
            dmi_display = pd.read_csv(largest_dmi_display_path)
            
        dmi_display.columns = dmi_display.columns.str.strip()
        dmi_display = dmi_display[['TrainDirection', 'TagNo', 'AbsLocation', 'TimeStamp']]
        dmi_display['TrainDirection'] = dmi_display['TrainDirection'].replace(
            {2: "Reverse", 1: "Nominal", 0: "Undefined"}
        )

        dmi_display_main = dmi_display.copy()
        dmi_display = dmi_display[dmi_display['TrainDirection'] != 'Undefined']

        direction_list = dmi_display['TrainDirection'].unique().tolist()

        for ref_dir in direction_list:

            print(f"Checking direction: {ref_dir}")

            try:

                master_data_set_path = os.path.join(master_data, f"{ref_dir}.xlsx")
                master_data_set = pd.read_excel(master_data_set_path)
            except:
                master_data_set_path = os.path.join(master_data, f"{ref_dir}.csv")
                master_data_set = pd.read_csv(master_data_set_path)


            dmi_display_dir = dmi_display[
                dmi_display['TrainDirection'] == ref_dir
            ].copy()

            try:
                report_generation(
                    master_data_set,
                    dmi_display_dir,
                    tag_read,
                    ref_dir,
                    main_fodler,
                    output,
                    division,
                    dmi_display_main
                )

            except Exception as e:
                print(f"Error while generating {stn}, {ref_dir}: {e}")
